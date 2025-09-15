from flask import Flask, render_template, request, send_file
from openpyxl import load_workbook
import datetime
import os

app = Flask(__name__)

@app.route("/")
def index():
    return render_template("index.html")

@app.route("/generar", methods=["POST"])
def generar():
    tipo = request.form["tipo"]
    ubicacion = request.form["ubicacion"]
    metros = float(request.form["metros"])
    precio = float(request.form["precio"])
    colocacion = float(request.form["colocacion"])
    flete = float(request.form["flete"])

    # Cargar plantilla
    wb = load_workbook("JDN_Plantilla_Cesped.xlsx")
    ws = wb.active

    # Escribir datos (asegurate que estas celdas no sean merged)
    ws["C2"] = tipo
    ws["C11"] = tipo
    ws["C4"] = ubicacion
    ws["G11"] = precio
    ws["G12"] = colocacion
    ws["G13"] = flete
    ws["H5"] = datetime.datetime.now()

    # Guardar archivo temporal
    filename = f"presupuesto_{ubicacion}_{tipo}.xlsx"
    wb.save(filename)

    # Enviar archivo como descarga
    return send_file(filename, as_attachment=True)

if __name__ == "__main__":
    app.run(host="0.0.0.0", port=5000, debug=True)

