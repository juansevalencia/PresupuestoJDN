import tkinter as tk
from tkinter import simpledialog, messagebox
from openpyxl import Workbook
import datetime
from openpyxl import load_workbook

def mostrar_menu_principal():
    limpiar_ventana()
    label = tk.Label(root, text="Seleccione Presupuesto", font=("Arial", 20))
    label.pack(pady=20)

    btn1 = tk.Button(root, text="Venta De Pasto (Sin Riego)", width=20, height=2, command= mostrar_submenu_pasto)
    btn1.pack(pady=10)

    ##btn2 = tk.Button(root, text="Movimiento De Suelos ", width=20, height=2, command=mostrar_submenu_suelos)
    ##btn2.pack(pady=10)
##
    ##btn3 = tk.Button(root, text="Parquizacion", width=20, height=2, command=mostrar_submenu_muelles)
    ##btn3.pack(pady=10)

def mostrar_submenu_pasto():
    limpiar_ventana()
    label = tk.Label(root, text = "Seleccione un tipo de presupuesto", font =("Arial", 20))
    label.pack(pady = 20)

    btn1 = tk.Button(root, text="Gramma Bahiana", width=30, height=2, command=lambda: generar_presupuesto("Gramma Bahiana"))
    btn1.pack(pady=10)

    btn2 = tk.Button(root, text="Bermuda", width=30, height=2, command=lambda: generar_presupuesto("Bermuda"))
    btn2.pack(pady=10)

    btn3 = tk.Button(root, text="Tiffway 419", width=30, height=2, command=lambda: generar_presupuesto("Tiffway 419"))
    btn3.pack(pady=10)

def limpiar_ventana():
    for widget in root.winfo_children():
        widget.destroy()


def generar_presupuesto(tipo):
    ubicacion = simpledialog.askstring("Ubicacion", "Ubicado en:")
    metrosCuadrados = float(simpledialog.askstring("Metros Cuadrados", "Metros cuadrados de Gramma Bahiana:"))
    precioPorMetroCuadrado = float(simpledialog.askstring("Precio Pasto m2" , "Precio por m2 de Gramma"))
    colocacion = float(simpledialog.askstring("Colocación","Si no incluye colocacion ponga el numero 0, caso contrario el precio de colocacion por m2"))
    flete = float(simpledialog.askstring("Flete","Si no incluye flete ponga el numero 0, caso contrario el precio:"))

    if not ubicacion or not metrosCuadrados or not precioPorMetroCuadrado:
        messagebox.showwarning("Error", "Debes completar todos los campos")
        return

    wb = load_workbook("JDN_Plantilla_Cesped.xlsx")
    ws = wb.active  # o wb["Hoja1"] si querés una hoja específica
    
    # Escribir datos en las celdas correctas
    ws["A2"] = tipo
    ws["A3"] = ubicacion
    ws["H5"] = datetime.datetime.now()
    ws["B11"]= metrosCuadrados
    ws["B12"] = metrosCuadrados
    ws["C11"] = tipo
    ws["G11"] = precioPorMetroCuadrado
    ws["G12"] = colocacion
    ws["G13"] = flete

    filename = f"presupuesto_{ubicacion}_{tipo}.xlsx"
    wb.save(filename)

    messagebox.showinfo("Exito", f"Presupuesto generado: {filename}")

root = tk.Tk()
root.title("Menú de Presupuestos")
root.geometry("800x600")
root.configure(bg ="#228B22")

mostrar_menu_principal()

root.mainloop()